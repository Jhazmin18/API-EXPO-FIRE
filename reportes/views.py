from io import BytesIO
from datetime import datetime, timedelta

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import legal, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from empresas.models import Contacto, Empresa
from empresas.serializers import ContactoSerializer, EmpresaSerializer
from extintores.models import Extintor
from usuarios.models import Perfil


ESTADO_PRIORIDAD = {
    'rojo': 0,
    'amarillo': 1,
    'verde': 2,
}

ESTADO_FILL = {
    'rojo': PatternFill('solid', fgColor='FCA5A5'),
    'amarillo': PatternFill('solid', fgColor='FEF08A'),
    'verde': PatternFill('solid', fgColor='BBF7D0'),
}

BRAND_ORANGE = colors.HexColor('#FF8A00')
BRAND_DARK = colors.HexColor('#1F2937')
BRAND_MUTED = colors.HexColor('#6B7280')
HEADER_FILL = PatternFill('solid', fgColor='E5E7EB')
TITLE_FILL = PatternFill('solid', fgColor='1F2937')


class InventarioReporteView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        empresa_id = request.query_params.get('empresa_id')
        if not empresa_id:
            return Response({'detail': 'Debe enviar empresa_id.'}, status=400)

        empresa = get_object_or_404(
            Empresa.objects.prefetch_related('contactos'),
            id=empresa_id,
        )
        primer_contacto = self._get_primer_contacto(empresa)
        extintores = list(
            Extintor.objects.filter(empresa=empresa)
            .select_related('empresa')
            .order_by('codigo')
        )

        resumen = {
            'total': len(extintores),
            'verde': 0,
            'amarillo': 0,
            'rojo': 0,
        }

        inventario = []
        for extintor in extintores:
            estado = extintor.estado
            resumen[estado] += 1
            inventario.append(self._serialize_extintor(extintor))

        inventario.sort(
            key=lambda item: (
                ESTADO_PRIORIDAD.get(item['estado'], 99),
                item['dias_para_revision'] if item['dias_para_revision'] is not None else 999999,
                item['dias_para_vencer'] if item['dias_para_vencer'] is not None else 999999,
                item['codigo'],
            )
        )
        alertas = [
            item for item in inventario
            if item['estado'] in {'rojo', 'amarillo'}
        ]

        data = {
            'fecha_generacion': timezone.localdate().isoformat(),
            'archivo_sugerido': self._build_filename(empresa),
            'empresa': EmpresaSerializer(empresa, context={'request': request}).data,
            'primer_contacto': (
                ContactoSerializer(primer_contacto).data
                if primer_contacto else None
            ),
            'resumen_estados': resumen,
            'inventario': inventario,
            'alertas': alertas,
        }

        formato = (
            request.query_params.get('formato')
            or request.query_params.get('format')
            or ''
        ).strip().lower()
        if formato == 'xlsx':
            return self._excel_response(data)
        if formato == 'pdf':
            return self._pdf_response(data)

        return Response(data)

    def _serialize_extintor(self, extintor):
        return {
            'id': str(extintor.id),
            'codigo': extintor.codigo,
            'ubicacion': extintor.ubicacion,
            'agente': extintor.tipo,
            'agente_display': extintor.get_tipo_display(),
            'capacidad': extintor.capacidad,
            'fecha_fabricacion': self._date(extintor.fecha_fabricacion),
            'fecha_vencimiento': self._date(extintor.fecha_vencimiento),
            'proxima_revision': self._date(extintor.proxima_revision),
            'fecha_prueba_hidrostatica': self._date(extintor.fecha_prueba_hidrostatica),
            'dias_para_vencer': extintor.dias_para_vencer,
            'dias_para_revision': extintor.dias_para_revision,
            'estado': extintor.estado,
            'observaciones': extintor.observaciones,
        }

    def _date(self, value):
        return value.isoformat() if value else None

    def _get_primer_contacto(self, empresa):
        return Contacto.objects.filter(empresa_id=empresa.id).order_by('created_at', 'id').first()

    def _build_filename(self, empresa):
        fecha = timezone.localdate().isoformat()
        nombre = ''.join(
            char if char.isalnum() else '_'
            for char in empresa.nombre.strip()
        ).strip('_')
        return f'Reporte_Extintores_{nombre}_{fecha}.xlsx'

    def _build_pdf_filename(self, xlsx_filename):
        if xlsx_filename.endswith('.xlsx'):
            return f'{xlsx_filename[:-5]}.pdf'
        return f'{xlsx_filename}.pdf'

    def _excel_response(self, data):
        workbook = Workbook()
        resumen_sheet = workbook.active
        resumen_sheet.title = 'Resumen'
        inventario_sheet = workbook.create_sheet('Inventario completo')
        alertas_sheet = workbook.create_sheet('Alertas')

        self._write_resumen(resumen_sheet, data)
        self._write_extintores(inventario_sheet, data['inventario'])
        self._write_extintores(alertas_sheet, data['alertas'])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{data["archivo_sugerido"]}"'
        return response

    def _write_resumen(self, sheet, data):
        empresa = data['empresa']
        contacto = data.get('primer_contacto') or {}
        resumen = data['resumen_estados']

        sheet.merge_cells('A1:D1')
        sheet['A1'] = 'Reporte de Inventario de Extintores'
        sheet['A1'].font = Font(bold=True, color='FFFFFF', size=14)
        sheet['A1'].fill = TITLE_FILL
        sheet['A1'].alignment = Alignment(horizontal='center')

        rows = [
            ('Empresa', empresa.get('nombre')),
            ('Razón social', empresa.get('razon_social')),
            ('Tipo de inmueble', empresa.get('tipo_inmueble')),
            ('Fecha de generación', data['fecha_generacion']),
            ('Primer contacto', contacto.get('nombre')),
            ('Correo principal', contacto.get('correo_principal')),
            ('Teléfono principal', contacto.get('telefono_principal')),
            ('', ''),
            ('Total extintores', resumen['total']),
            ('Verde', resumen['verde']),
            ('Amarillo', resumen['amarillo']),
            ('Rojo', resumen['rojo']),
        ]

        for row_number, row in enumerate(rows, start=3):
            sheet.cell(row=row_number, column=1, value=row[0])
            value = row[1] if row[1] is not None else ''
            sheet.cell(row=row_number, column=2, value=value)
            sheet.cell(row=row_number, column=1).font = Font(bold=True)

        sheet.column_dimensions['A'].width = 24
        sheet.column_dimensions['B'].width = 44

    def _write_extintores(self, sheet, rows):
        columns = [
            ('Código', 'codigo'),
            ('Ubicación', 'ubicacion'),
            ('Agente', 'agente_display'),
            ('Capacidad', 'capacidad'),
            ('Fecha Fabricación', 'fecha_fabricacion'),
            ('Fecha Vencimiento', 'fecha_vencimiento'),
            ('Próxima Revisión', 'proxima_revision'),
            ('Prueba Hidrostática', 'fecha_prueba_hidrostatica'),
            ('Días para Vencer', 'dias_para_vencer'),
            ('Días para Revisión', 'dias_para_revision'),
            ('Estado', 'estado'),
            ('Observaciones', 'observaciones'),
        ]

        for column_number, (header, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=column_number, value=header)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        for row_number, item in enumerate(rows, start=2):
            for column_number, (_, key) in enumerate(columns, start=1):
                cell = sheet.cell(row=row_number, column=column_number, value=item.get(key))
                cell.alignment = Alignment(vertical='top', wrap_text=True)

            estado_cell = sheet.cell(row=row_number, column=11)
            estado_cell.fill = ESTADO_FILL.get(item.get('estado'), HEADER_FILL)
            estado_cell.font = Font(bold=True)

        widths = [16, 32, 26, 14, 18, 18, 18, 20, 18, 20, 14, 42]
        for column_number, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_number)].width = width

        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions

    def _pdf_response(self, data):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(legal),
            rightMargin=0.35 * inch,
            leftMargin=0.35 * inch,
            topMargin=0.62 * inch,
            bottomMargin=0.5 * inch,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            alignment=TA_CENTER,
            fontSize=16,
            leading=20,
            textColor=colors.HexColor('#1F2937'),
        )
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=11,
            leading=14,
            textColor=colors.HexColor('#1F2937'),
            spaceBefore=8,
            spaceAfter=6,
        )
        small_style = ParagraphStyle(
            'Small',
            parent=styles['BodyText'],
            fontSize=7,
            leading=9,
        )

        story = [
            Paragraph('Reporte de Inventario de Extintores', title_style),
            Spacer(1, 8),
        ]
        story.extend(self._pdf_resumen_table(data))
        story.extend([
            PageBreak(),
            self._pdf_page_header(data),
            Spacer(1, 6),
            Paragraph('Inventario completo', section_style),
            self._pdf_extintores_table(data['inventario'], small_style),
            PageBreak(),
            self._pdf_page_header(data),
            Spacer(1, 6),
            Paragraph('Alertas', section_style),
            self._pdf_extintores_table(data['alertas'], small_style),
        ])

        doc.build(
            story,
            onFirstPage=self._draw_pdf_branding,
            onLaterPages=self._draw_pdf_branding,
        )
        buffer.seek(0)

        filename = self._build_pdf_filename(data['archivo_sugerido'])
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _draw_pdf_branding(self, canvas, doc):
        canvas.saveState()
        page_width, page_height = landscape(legal)

        canvas.setFillColor(BRAND_ORANGE)
        canvas.rect(0, page_height - 0.18 * inch, page_width, 0.05 * inch, stroke=0, fill=1)

        canvas.setFillColor(BRAND_DARK)
        canvas.setFont('Helvetica-Bold', 9)
        canvas.drawString(doc.leftMargin, page_height - 0.38 * inch, 'EXPRO FIRE')

        canvas.setFillColor(BRAND_MUTED)
        canvas.setFont('Helvetica', 7)
        canvas.drawRightString(
            page_width - doc.rightMargin,
            page_height - 0.38 * inch,
            'Sistema de gestion de extintores',
        )

        canvas.setStrokeColor(colors.HexColor('#E5E7EB'))
        canvas.setLineWidth(0.4)
        canvas.line(doc.leftMargin, 0.32 * inch, page_width - doc.rightMargin, 0.32 * inch)

        canvas.setFillColor(BRAND_MUTED)
        canvas.setFont('Helvetica', 7)
        canvas.drawString(doc.leftMargin, 0.18 * inch, 'EXPRO FIRE')
        canvas.drawRightString(
            page_width - doc.rightMargin,
            0.18 * inch,
            f'Pagina {doc.page}',
        )
        canvas.restoreState()

    def _pdf_page_header(self, data):
        empresa = data['empresa']
        rows = [[
            f"Empresa: {empresa.get('nombre') or ''}",
            f"Fecha de generación: {data['fecha_generacion']}",
        ]]
        table = Table(rows, colWidths=[5.7 * inch, 4.6 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        return table

    def _pdf_resumen_table(self, data):
        empresa = data['empresa']
        contacto = data.get('primer_contacto') or {}
        resumen = data['resumen_estados']
        rows = [
            ['Empresa', empresa.get('nombre') or '', 'Fecha', data['fecha_generacion']],
            ['Razon social', empresa.get('razon_social') or '', 'Tipo inmueble', empresa.get('tipo_inmueble') or ''],
            ['Primer contacto', contacto.get('nombre') or '', 'Correo', contacto.get('correo_principal') or ''],
            ['Telefono', contacto.get('telefono_principal') or '', 'Total extintores', resumen['total']],
        ]
        table = Table(rows, colWidths=[1.2 * inch, 3.0 * inch, 1.35 * inch, 2.8 * inch])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D1D5DB')),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E5E7EB')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        return [
            table,
            Spacer(1, 12),
            self._pdf_estado_cards(resumen),
        ]

    def _pdf_estado_cards(self, resumen):
        cards = [
            ['Verde', resumen['verde'], colors.HexColor('#DCFCE7'), colors.HexColor('#15803D')],
            ['Amarillo', resumen['amarillo'], colors.HexColor('#FEF9C3'), colors.HexColor('#A16207')],
            ['Rojo', resumen['rojo'], colors.HexColor('#FEE2E2'), colors.HexColor('#B91C1C')],
        ]
        table_data = [[
            Paragraph(f'<b>{label}</b><br/><font size="18">{count}</font>', ParagraphStyle(
                f'Estado{label}',
                alignment=TA_CENTER,
                fontSize=9,
                leading=14,
                textColor=text_color,
            ))
            for label, count, _, text_color in cards
        ]]
        table = Table(table_data, colWidths=[1.45 * inch, 1.45 * inch, 1.45 * inch], hAlign='CENTER')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), cards[0][2]),
            ('BACKGROUND', (1, 0), (1, 0), cards[1][2]),
            ('BACKGROUND', (2, 0), (2, 0), cards[2][2]),
            ('BOX', (0, 0), (0, 0), 0.6, colors.HexColor('#BBF7D0')),
            ('BOX', (1, 0), (1, 0), 0.6, colors.HexColor('#FDE68A')),
            ('BOX', (2, 0), (2, 0), 0.6, colors.HexColor('#FCA5A5')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        return table

    def _pdf_extintores_table(self, rows, small_style):
        headers = [
            'Codigo',
            'Ubicacion',
            'Agente',
            'Cap.',
            'Fabricacion',
            'Vencimiento',
            'Prox. rev.',
            'Prueba hidr.',
            'Dias venc.',
            'Dias rev.',
            'Estado',
            'Observaciones',
        ]
        table_data = [headers]
        for item in rows:
            table_data.append([
                self._p(item.get('codigo'), small_style),
                self._p(item.get('ubicacion'), small_style),
                self._p(item.get('agente_display'), small_style),
                self._p(item.get('capacidad'), small_style),
                self._p(item.get('fecha_fabricacion'), small_style),
                self._p(item.get('fecha_vencimiento'), small_style),
                self._p(item.get('proxima_revision'), small_style),
                self._p(item.get('fecha_prueba_hidrostatica'), small_style),
                item.get('dias_para_vencer'),
                item.get('dias_para_revision'),
                self._p(item.get('estado'), small_style),
                self._p(item.get('observaciones'), small_style),
            ])

        if len(table_data) == 1:
            table_data.append(['Sin datos', '', '', '', '', '', '', '', '', '', '', ''])

        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[
                0.75 * inch,
                1.35 * inch,
                1.25 * inch,
                0.55 * inch,
                0.78 * inch,
                0.85 * inch,
                0.85 * inch,
                0.78 * inch,
                0.65 * inch,
                0.65 * inch,
                0.6 * inch,
                1.65 * inch,
            ],
        )
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('BACKGROUND', (10, 1), (10, -1), colors.HexColor('#F3F4F6')),
        ]))
        return table

    def _p(self, value, style):
        return Paragraph(str(value or ''), style)


class LogUsuariosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        queryset = Perfil.objects.select_related('user', 'empresa').all()
        empresa_id = request.query_params.get('empresa_id')
        usuario_id = request.query_params.get('usuario_id')

        if empresa_id:
            queryset = queryset.filter(empresa_id=empresa_id)
        if usuario_id:
            queryset = queryset.filter(user_id=usuario_id)

        eventos = []
        for perfil in queryset:
            usuario = perfil.user
            usuario_data = self._usuario_data(perfil)

            eventos.append({
                'id': f'usuario_creado_{usuario.id}',
                'tipo': 'USUARIO',
                'accion': 'CREACION',
                'fecha': usuario.date_joined.isoformat() if usuario.date_joined else None,
                'descripcion': f'Usuario creado: {usuario.username}',
                'usuario': usuario_data,
            })

            eventos.append({
                'id': f'perfil_actualizado_{perfil.id}',
                'tipo': 'USUARIO',
                'accion': 'ACTUALIZACION_PERFIL',
                'fecha': perfil.updated_at.isoformat() if perfil.updated_at else None,
                'descripcion': f'Perfil actualizado: {usuario.username}',
                'usuario': usuario_data,
            })

            if usuario.last_login:
                eventos.append({
                    'id': f'usuario_login_{usuario.id}',
                    'tipo': 'USUARIO',
                    'accion': 'ULTIMO_LOGIN',
                    'fecha': usuario.last_login.isoformat(),
                    'descripcion': f'Ultimo acceso: {usuario.username}',
                    'usuario': usuario_data,
                })

        eventos = self._filtrar_eventos(request, eventos)
        return Response({
            'total': len(eventos),
            'resultados': eventos,
        })

    def _usuario_data(self, perfil):
        usuario = perfil.user
        return {
            'id': usuario.id,
            'username': usuario.username,
            'email': usuario.email,
            'nombre_completo': perfil.nombre_completo,
            'rol': perfil.rol,
            'rol_display': perfil.get_rol_display(),
            'empresa_id': perfil.empresa_id,
            'empresa_nombre': perfil.empresa.nombre if perfil.empresa else None,
            'activo': usuario.is_active,
        }

    def _filtrar_eventos(self, request, eventos):
        dias = _parse_int(request.query_params.get('dias'))
        limite = _parse_int(request.query_params.get('limite'), default=100)

        if dias:
            desde = timezone.now() - timedelta(days=dias)
            eventos = [
                evento for evento in eventos
                if evento['fecha'] and _parse_datetime(evento['fecha']) >= desde
            ]

        eventos = sorted(
            eventos,
            key=lambda evento: evento['fecha'] or '',
            reverse=True,
        )
        return eventos[:limite]


class LogExtintoresView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        queryset = Extintor.objects.select_related('empresa', 'creado_por').all()
        empresa_id = request.query_params.get('empresa_id')
        extintor_id = request.query_params.get('extintor_id')
        creado_por = request.query_params.get('creado_por')
        estado = request.query_params.get('estado')

        if empresa_id:
            queryset = queryset.filter(empresa_id=empresa_id)
        if extintor_id:
            queryset = queryset.filter(id=extintor_id)
        if creado_por:
            queryset = queryset.filter(creado_por_id=creado_por)

        eventos = []
        for extintor in queryset:
            if estado and extintor.estado != estado:
                continue

            extintor_data = self._extintor_data(extintor)
            creador = self._usuario_basico(extintor.creado_por)

            eventos.append({
                'id': f'extintor_creado_{extintor.id}',
                'tipo': 'EXTINTOR',
                'accion': 'CREACION',
                'fecha': extintor.created_at.isoformat() if extintor.created_at else None,
                'descripcion': f'Extintor creado: {extintor.codigo}',
                'usuario': creador,
                'extintor': extintor_data,
            })

            eventos.append({
                'id': f'extintor_actualizado_{extintor.id}',
                'tipo': 'EXTINTOR',
                'accion': 'ACTUALIZACION',
                'fecha': extintor.updated_at.isoformat() if extintor.updated_at else None,
                'descripcion': f'Extintor actualizado: {extintor.codigo}',
                'usuario': creador,
                'extintor': extintor_data,
            })

        eventos = self._filtrar_eventos(request, eventos)
        return Response({
            'total': len(eventos),
            'resultados': eventos,
        })

    def _extintor_data(self, extintor):
        return {
            'id': str(extintor.id),
            'codigo': extintor.codigo,
            'ubicacion': extintor.ubicacion,
            'tipo': extintor.tipo,
            'tipo_display': extintor.get_tipo_display(),
            'capacidad': extintor.capacidad,
            'empresa_id': extintor.empresa_id,
            'empresa_nombre': extintor.empresa.nombre if extintor.empresa else None,
            'estado': extintor.estado,
            'dias_para_vencer': extintor.dias_para_vencer,
            'dias_para_revision': extintor.dias_para_revision,
        }

    def _usuario_basico(self, usuario):
        if not usuario:
            return None
        return {
            'id': usuario.id,
            'username': usuario.username,
            'email': usuario.email,
            'nombre': usuario.get_full_name() or usuario.username,
        }

    def _filtrar_eventos(self, request, eventos):
        dias = _parse_int(request.query_params.get('dias'))
        limite = _parse_int(request.query_params.get('limite'), default=100)

        if dias:
            desde = timezone.now() - timedelta(days=dias)
            eventos = [
                evento for evento in eventos
                if evento['fecha'] and _parse_datetime(evento['fecha']) >= desde
            ]

        eventos = sorted(
            eventos,
            key=lambda evento: evento['fecha'] or '',
            reverse=True,
        )
        return eventos[:limite]


def _parse_int(value, default=None):
    if value in (None, ''):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _parse_datetime(value):
    parsed = datetime.fromisoformat(value)
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed)
    return parsed
